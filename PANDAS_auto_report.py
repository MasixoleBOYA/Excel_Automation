import pandas as pd
from openpyxl import Workbook, load_workbook

work_book = load_workbook('C:/Users/J1121857/Downloads/GANTRY_RAW_data.xlsx')
product_codes_workbook = load_workbook('C:/Users/J1121857/Downloads/Copy of Reseller customer list 29 Mar 22.XLSX')  # Replace with the actual path
customer_codes_workbook = load_workbook('C:/Users/J1121857/Downloads/Reseller ship-to list.xlsx')  # Replace with the actual path

print(f"\nMAIN WORKBOOK sheetnames:\n{work_book.sheetnames}")
print(f"\nPRODUCT workbook sheetnames:\n{product_codes_workbook.sheetnames}")
print(f"\nCUSTOMER CODES worksheet sheetnames:\n{customer_codes_workbook.sheetnames}")



depots = ["Alrode", "Bethlehem", "Cape Town", "East London", "Island View", "Klerksdorp", "Ladysmith", "Mossel Bay",
          "Nelspruit", "Port Elizabeth", "Sasolburg", "Tarlton", "Waltloo", "Witbank"]

# PART 1: renames the "Gantry AP" column for all sheets
def sheet_rename() -> None:
    for i in work_book.sheetnames:
        work_sheet = work_book[i]
        column_index = 1

        for x in work_sheet.iter_rows(min_row=2, max_row=work_sheet.max_row, min_col=column_index, max_col=column_index):
            for cell in x:
                cell.value = i

# PART 2: Append data from other sheets to a specified sheet
def appending_to_onesheet(sheet_name: str) -> None:
    appending_sheet = work_book[sheet_name]

    for depot_name in work_book.sheetnames[1:]:
        current_worksheet = work_book[depot_name]
        for row in current_worksheet.iter_rows(min_row=2, max_row=current_worksheet.max_row, values_only=True):
            appending_sheet.append(row)

# PART 3: Add "Customer Names" column to Alrode sheet using pandas
def add_customer_names_column():
    alrode_sheet = work_book["Alrode"]

    alrode_df = pd.DataFrame(alrode_sheet.values, columns=[col[0].value for col in alrode_sheet.iter_cols()])
    customer_codes_workbook_sheet = customer_codes_workbook["Cust Loc (3)"]
    customer_codes_workbook_df = pd.DataFrame(customer_codes_workbook_sheet.values, columns=[col[0].value for col in customer_codes_workbook_sheet.iter_cols()] )

    #Look for customer names based on the "Customer Codes" column
    alrode_df["Customer Names"] = alrode_df["Customer No."].map(customer_codes_workbook_df.set_index('Customer No')["Customer Name"])
    '''
    AttributeError: 'Workbook' object has no attribute 'set_index'. Did you mean: 'get_index'?
PS C:\Users\J1121857> & C:/Users/J1121857/AppData/Local/Programs/Python/Python312/python.exe c:/Users/J1121857/Downloads/PANDAS_auto_report.py

MAIN WORKBOOK sheetnames:
['Alrode', 'Bethlehem', 'Cape Town', 'East London', 'Island View ', 'Klerksdorp', 'Ladysmith', 'Mossel Bay', 'Nelspruit', 'Port 
Elizabeth', 'Sasolburg', 'Tarlton ', 'Waltloo', 'Witbank']

PRODUCT workbook sheetnames:
['sold to and ship to', 'Active credit accounts', 'Sold to', 'SAP shortcuts', 'Product Code']

CUSTOMER CODES worksheet sheetnames:
['Depots', 'Sheet2', 'Recoveries', 'Channel', 'Primary Tpt', 'Depot Cost', 'Cust Loc (3)', 'Cust Loc (2)', 'Customer Info (2)', 
'Customer Info', 'Cust Loc', 'Sheet1']

WORKBOOK sheetnames:
['Alrode', 'Bethlehem', 'Cape Town', 'East London', 'Island View ', 'Klerksdorp', 'Ladysmith', 'Mossel Bay', 'Nelspruit', 'Port 
Elizabeth', 'Sasolburg', 'Tarlton ', 'Waltloo', 'Witbank']
Traceback (most recent call last):
  File "c:\Users\J1121857\Downloads\PANDAS_auto_report.py", line 77, in <module>
    add_customer_names_column()
  File "c:\Users\J1121857\Downloads\PANDAS_auto_report.py", line 45, in add_customer_names_column
    alrode_df["Customer Names"] = alrode_df["Customer No."].map(customer_codes_workbook_df.set_index('Customer No')["Customer Name"])
                                  ^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
  File "C:\Users\J1121857\AppData\Local\Programs\Python\Python312\Lib\site-packages\pandas\core\series.py", line 4691, in map   
    new_values = self._map_values(arg, na_action=na_action)
                 ^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
  File "C:\Users\J1121857\AppData\Local\Programs\Python\Python312\Lib\site-packages\pandas\core\base.py", line 921, in _map_values
    return algorithms.map_array(arr, mapper, na_action=na_action, convert=convert)
           ^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
  File "C:\Users\J1121857\AppData\Local\Programs\Python\Python312\Lib\site-packages\pandas\core\algorithms.py", line 1732, in map_array
    indexer = mapper.index.get_indexer(arr)
              ^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
  File "C:\Users\J1121857\AppData\Local\Programs\Python\Python312\Lib\site-packages\pandas\core\indexes\base.py", line 3885, in 
get_indexer
    raise InvalidIndexError(self._requires_unique_msg)
pandas.errors.InvalidIndexError: Reindexing only valid with uniquely valued Index objects
PS C:\Users\J1121857> 
    '''

    alrode_df["Customer Names"].fillna("", inplace=True)

    #Update the Alrode sheet, considering the new column
    alrode_sheet.clear()
    alrode_sheet.append(list(alrode_df.columns))
    for row in alrode_df.values:
        alrode_sheet.append(list(row))

# PART 4: Add "Product Names" column to Alrode sheet using pandas
def add_product_names_column():
    alrode_sheet = work_book["Alrode"]
    product_codes_workbook = product_codes_workbook['Product Code']

    alrode_df = pd.DataFrame(alrode_sheet.values, columns=[col[0].value for col in alrode_sheet.iter_cols()])

    # Looking up for product names based on the "Product" column
    alrode_df["Product Names"] = alrode_df["Product"].map(product_codes_workbook.set_index("Product code")["Product name"])

    alrode_df["Product Names"].fillna("", inplace=True)

    # Update the Alrode sheet 
    alrode_sheet.clear()
    alrode_sheet.append(list(alrode_df.columns))
    for row in alrode_df.values:
        alrode_sheet.append(list(row))

print(f"\nWORKBOOK sheetnames:\n{work_book.sheetnames}")
sheet_rename()
appending_to_onesheet("Alrode")

add_customer_names_column()
add_product_names_column()

work_book.save('C:/Users/J1121857/Downloads/AAAAAAAAAAAA.xlsx')
