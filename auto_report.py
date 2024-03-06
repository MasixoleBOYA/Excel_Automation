from openpyxl import Workbook, load_workbook


work_book = load_workbook('C:/Users/J1121857/Downloads/GANTRY_RAW_data.xlsx') 
product_codes_workbook = load_workbook('path_to_product_codes_workbook.xlsx')  # Replace with the actual path
customer_codes_workbook = load_workbook('path_to_customer_codes_workbook.xlsx')  # Replace with the actual path


depots = ["Alrode",
          "Bethlehem",
          "Cape Town",
          "East London",
          "Island View",
          "Klerksdorp",
          "Ladysmith",
          "Mossel Bay",
          "Nelspruit",
          "Port Elizabeth",
          "Sasolburg",
          "Tarlton",
          "Waltloo",
          "Witbank"]


# PART 1: renames the "Gantry AP" column for all sheets
def sheet_rename()->None:
    for i in work_book.sheetnames: 
        work_sheet = work_book[i] #for each depot, consider the sheet that matches the depot
        column_index = 1

        for x in work_sheet.iter_rows(min_row=2, max_row=work_sheet.max_row, min_col=column_index, max_col=column_index):
            for cell in x:
                cell.value = i



        # for col in work_sheet.iter_cols(min_row=1,max_row=1,min_col=1):
            # print(col[0].value)
            # print(col[0].column)
            # if col[0].value == "Gantry AP":    
            #     column_index = col[0].column
            #     working_column = col[0].value
            #     # break
                     
            # print("x1")
            # column = work_sheet[column_index]
            # print("x2")



            


            # for j in column:
            #     # print(f"VALUES OF COLUMN ARE: \n{j.value}")
            #     j.value = i
            # print("x3")

          

# SECOND PART
def appending_to_onesheet(sheetName: str) -> None:
    appending_sheet = work_book[sheetName]
          
    for depot_name in depots[1:]:
        current_worksheet = work_book[depot_name]
        for row in current_worksheet.iter_rows(min_row =2,
                                                     max_row = current_worksheet.max_row,
                                                     values_only = True):
            appending_sheet.append(row)


# XXXXXXXXXXXXXXXXXXXXX proposed using openpyxl XXXXXXXXXXXXXXXXXXXXX
            
# PART 3: Add "Customer Names" column to Alrode sheet
def add_customer_names_column():
    alrode_sheet = work_book["Alrode"]
    customer_codes_sheet = customer_codes_workbook.active

    # Insert a new column after the "Customer Codes" column
    alrode_sheet.insert_cols(3)

    # Iterate through rows in the Alrode sheet
    for row in alrode_sheet.iter_rows(min_row=2, max_row=alrode_sheet.max_row, min_col=3, max_col=3):
        for cell in row:
            customer_code = cell.value
            if customer_code is not None:
                # Find the corresponding customer name in the customer codes workbook
                customer_name = customer_codes_sheet[f'B{customer_code}'].value
                # Insert the customer name in the new column
                alrode_sheet.cell(row=cell.row, column=cell.column + 1, value=customer_name)

# PART 4: Add "Product Names" column to Alrode sheet
def add_product_names_column():
    alrode_sheet = work_book["Alrode"]
    product_codes_sheet = product_codes_workbook.active

    # Insert a new column after the "Product" column
    alrode_sheet.insert_cols(5)

    # Iterate through rows in the Alrode sheet
    for row in alrode_sheet.iter_rows(min_row=2, max_row=alrode_sheet.max_row, min_col=5, max_col=5):
        for cell in row:
            product_code = cell.value
            if product_code is not None:
                # Find the corresponding product name in the product codes workbook
                product_name = product_codes_sheet[f'B{product_code}'].value
                # Insert the product name in the new column
                alrode_sheet.cell(row=cell.row, column=cell.column + 1, value=product_name)


if __name__ == "__main__":

    print(work_book.sheetnames)
    sheet_rename()
    appending_to_onesheet("Alrode")
    work_book.save('C:/Users/J1121857/Downloads/TEST_NEW_NEW.xlsx')
          
