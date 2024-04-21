import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.pivot_table import PivotTable, TableCache

from customer_codes_data import customer_codesNames_dictionary
from product_codes_data import product_codes_dictionary



work_book = load_workbook('C:/Users/J1121857/Downloads/GANTRY_RAW_data.xlsx')
product_codes_workbook = load_workbook('C:/Users/J1121857/Downloads/Copy of Reseller customer list 29 Mar 22.XLSX')  # Replace with the actual path
# customer_codes_workbook = load_workbook('C:/Users/J1121857/Downloads/Reseller ship-to list.xlsx')  # Replace with the actual path

depots = ["Alrode", "Bethlehem", "Cape Town", "East London", "Island View", "Klerksdorp", "Ladysmith", "Mossel Bay",
          "Nelspruit", "Port Elizabeth", "Sasolburg", "Tarlton", "Waltloo", "Witbank"]

# PART 1: renames the "Gantry AP" column for all sheets
def sheet_rename() -> None:
    for i in work_book.sheetnames:
        work_sheet = work_book[i]
        column_index = 1

        for x in work_sheet.iter_rows(min_row=2, max_row=work_sheet.max_row, min_col=column_index, max_col=column_index):
            for cell in x:
                cell.value = i

# PART 2: Append data from other sheets to a specified sheet
def appending_to_onesheet(sheet_name: str) -> None:
    appending_sheet = work_book[sheet_name]

    for depot_name in work_book.sheetnames[1:]:
        current_worksheet = work_book[depot_name]
        print(f"\n Working on sheet : {current_worksheet}\n")
        for row in current_worksheet.iter_rows(min_row=2, max_row=current_worksheet.max_row, values_only=True):
            appending_sheet.append(row)

# PART 3: Add "Customer Names" column to Alrode sheet using pandas
def add_customer_names_column():
    alrode_sheet = work_book["Alrode"]
    
    alrode_df = pd.DataFrame(alrode_sheet.values, columns=[col[0].value for col in alrode_sheet.iter_cols()])
    
    # Convert 'Customer No.' column to numeric, coercing non-numeric values to NaN
    alrode_df['Customer No.'] = pd.to_numeric(alrode_df["Customer No."], errors='coerce')

    # Drop rows with NaN values in 'Customer No.' column
    alrode_df.dropna(subset=['Customer No.'], inplace=True)

    # Convert 'Customer No.' column to integers
    alrode_df['Customer No.'] = alrode_df['Customer No.'].astype(int)

    # Initialize an empty list to store customer names
    customer_names = []

    # Iterate over the rows of the DataFrame
    for index, row in alrode_df.iterrows():
        # Get the customer number from the current row
        customer_no = row['Customer No.']
        
        # Look up the customer name in the dictionary
        customer_name = customer_codesNames_dictionary.get(customer_no)
        
        # Append the customer name to the list
        customer_names.append(customer_name)

    # Insert a new column for 'Customer Names' at index 2 with custom heading
    alrode_sheet.insert_cols(2)
    alrode_sheet.cell(row=1, column=2, value="Customer Names")

    # Write the customer names into the new column
    for index, value in enumerate(customer_names, start=2):
        alrode_sheet.cell(row=index, column=2, value=value)

# PART 4: Add "Product Names" column to Alrode sheet using pandas
def add_product_names_column():
    alrode_sheet = work_book["Alrode"]
    
    alrode_df = pd.DataFrame(alrode_sheet.values, columns=[col[0].value for col in alrode_sheet.iter_cols()])
    
    # Convert 'Product' column to numeric, coercing non-numeric values to NaN
    alrode_df['Product'] = pd.to_numeric(alrode_df["Product"], errors='coerce')

    # Drop rows with NaN values in 'Product' column
    alrode_df.dropna(subset=['Product'], inplace=True)

    # Convert 'Product' column to integers
    alrode_df['Product'] = alrode_df['Product'].astype(int)

    # Initialize an empty list to store product names
    product_names = []

    # Iterate over the rows of the DataFrame
    for index, row in alrode_df.iterrows():
        # Get the product code from the current row
        product_code = row['Product']
        
        # Look up the product name in the dictionary
        product_name = product_codes_dictionary.get(product_code)
        
        # Append the product name to the list
        product_names.append(product_name)

    # Insert a new column for 'Product Names' at index 4 with custom heading
    alrode_sheet.insert_cols(4)
    alrode_sheet.cell(row=1, column=4, value="Product Names")

    # Write the product names into the new column
    for index, value in enumerate(product_names, start=2):
        alrode_sheet.cell(row=index, column=4, value=value)


# Additional functionality: Delete other sheets except "Alrode" sheet and remove rows with empty "Customer Names" column
def cleanup_workbook():
    sheets_to_keep = ["Alrode"]
    sheets_to_delete = [sheet for sheet in work_book.sheetnames if sheet not in sheets_to_keep]
    for sheet_name in sheets_to_delete:
        del work_book[sheet_name]
    alrode_sheet = work_book["Alrode"]
    customer_names_column = None
    for col in alrode_sheet.iter_cols(min_col=1, max_col=alrode_sheet.max_column, values_only=True):
        if col[0] == "Customer Names":
            customer_names_column = col
            break
    if customer_names_column is not None:
        rows_to_delete = []
        for i, cell_value in enumerate(customer_names_column[1:], start=2):
            if cell_value is None or cell_value == "":
                rows_to_delete.append(i)
        for row_index in reversed(rows_to_delete):
            alrode_sheet.delete_rows(row_index)

# Remove rows in "Alrode" sheet where "Product Names" column is null or empty
for col in alrode_sheet.iter_cols(min_col=1, max_col=alrode_sheet.max_column, values_only=True):
        if col[0] == "Product Names":
            product_names_column = col
            break
    product_names_column = None  
    rows_to_delete_again = []
    for x, product in enumerate(product_names_column[1:], start = 2):
        if product is None or product == "":
            rows_to_delete_again.append(x)

    for row_index in reversed(rows_to_delete_again):
        alrode_sheet.delete_rows(row_index)

# Create sheets for each pivot table
delivery_date_sheet = work_book.create_sheet("per delivery date")
cash_terms_sheet = work_book.create_sheet("cash terms")
depot_product_sheet = work_book.create_sheet("per depot per product")

def insert_blank_pivot_tables(work_book):
    # Create a new sheet for each pivot table
    delivery_date_sheet = work_book.create_sheet("per delivery date")
    cash_terms_sheet = work_book.create_sheet("cash terms")
    depot_product_sheet = work_book.create_sheet("per depot per product")
    
    # Get data from the "Alrode" sheet
    alrode_sheet = work_book["Alrode"]
    data_range = alrode_sheet.dimensions

    # Create PivotTable objects
    delivery_date_pivot = PivotTable()
    cash_terms_pivot = PivotTable()
    depot_product_pivot = PivotTable()

    # Set the range for each PivotTable
    delivery_date_pivot.range(ref=data_range)
    cash_terms_pivot.range(ref=data_range)
    depot_product_pivot.range(ref=data_range)

    # Add PivotTables to corresponding sheets
    delivery_date_sheet.add_pivot(delivery_date_pivot)
    cash_terms_sheet.add_pivot(cash_terms_pivot)
    depot_product_sheet.add_pivot(depot_product_pivot)



print(f"\nWORKBOOK sheetnames:\n{work_book.sheetnames}")
sheet_rename()
appending_to_onesheet("Alrode")

add_customer_names_column()
add_product_names_column()
cleanup_workbook()
insert_blank_pivot_tables(work_book)


print("cccccccccc DONE cccccccccccc")

work_book.save('C:/Users/J1121857/Downloads/AAAAAAAAAAAA.xlsx')
